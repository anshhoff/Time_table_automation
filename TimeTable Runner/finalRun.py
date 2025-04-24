import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def process_timetable(final_csv, courses_csv=None, institution_name="INDIAN INSTITUTE OF INFORMATION TECHNOLOGY", 
                     academic_session="Academic Session 2024-25"):
    # Read the timetable data
    df = pd.read_csv(final_csv)
    df.columns = df.columns.str.strip()
    
    # Process courses info if available
    courses_info = {}
    if courses_csv and os.path.exists(courses_csv):
        courses_df = pd.read_csv(courses_csv)
        for _, row in courses_df.iterrows():
            courses_info[row["Course"]] = {
                "name": row.get("Course_Name", ""),
                "credits": row.get("Credits", "3-0-0-3"),
                "faculty": row.get("Faculty", "")
            }
    
    # Function to parse time string to datetime
    def parse_time(time_str):
        return datetime.strptime(time_str, '%H:%M')
    
    # Detect all time points for each batch
    def get_time_points_for_batch(batch_df):
        time_points = set()
        
        for time_slot in batch_df["Time"].unique():
            try:
                start_str, end_str = time_slot.split('-')
                time_points.add(parse_time(start_str))
                time_points.add(parse_time(end_str))
            except:
                continue
                
        return sorted(time_points)
    
    # Generate all possible time slots for each batch
    def generate_time_slots(time_points):
        slots = []
        for i in range(len(time_points) - 1):
            start = time_points[i]
            end = time_points[i + 1]
            
            # Only include slots that are at least 15 minutes
            duration = (end - start).total_seconds() / 60
            if duration >= 15:
                slot = f"{start.strftime('%H:%M')}-{end.strftime('%H:%M')}"
                slots.append(slot)
        
        return slots
    
    # Add standard breaks to a specific batch
    def add_breaks_for_batch(batch_df, all_slots):
        # Define standard break times (approximate)
        morning_break_times = [(parse_time('10:30'), parse_time('10:45')),
                              (parse_time('10:45'), parse_time('11:00'))]
        lunch_break_times = [(parse_time('12:15'), parse_time('13:15')),
                            (parse_time('13:00'), parse_time('14:00'))]
        afternoon_break_times = [(parse_time('15:45'), parse_time('16:15')),
                                (parse_time('16:00'), parse_time('16:15'))]
        
        # Check if any slot overlaps with break times
        break_entries = []
        days = batch_df["Day"].unique()
        batch = batch_df["Batch"].iloc[0] if not batch_df.empty else "UNKNOWN"
        
        for day in days:
            day_df = batch_df[batch_df["Day"] == day]
            
            # Check for morning break
            morning_break_found = False
            for start, end in morning_break_times:
                for slot in all_slots:
                    slot_start, slot_end = slot.split('-')
                    slot_start = parse_time(slot_start)
                    slot_end = parse_time(slot_end)
                    
                    # If slot is entirely within break time & no classes overlap
                    if (start <= slot_start < end and slot_end <= end):
                        # Check if no classes overlap with this slot
                        has_overlap = False
                        for _, row in day_df.iterrows():
                            class_start, class_end = row["Time"].split('-')
                            class_start = parse_time(class_start)
                            class_end = parse_time(class_end)
                            
                            if (class_start < slot_end and class_end > slot_start):
                                has_overlap = True
                                break
                        
                        if not has_overlap:
                            break_entries.append({
                                "Day": day,
                                "Time": f"{slot_start.strftime('%H:%M')}-{slot_end.strftime('%H:%M')}",
                                "Room": "N/A",
                                "Batch": batch,
                                "Course": "Break",
                                "Type": "BREAK",
                                "Faculty": "N/A",
                                "Details": "Morning Break"
                            })
                            morning_break_found = True
                            break
                
                if morning_break_found:
                    break
            
            # Similar logic for lunch break
            lunch_break_found = False
            for start, end in lunch_break_times:
                for slot in all_slots:
                    slot_start, slot_end = slot.split('-')
                    slot_start = parse_time(slot_start)
                    slot_end = parse_time(slot_end)
                    
                    if (start <= slot_start < end and slot_end <= end):
                        # Check if no classes overlap
                        has_overlap = False
                        for _, row in day_df.iterrows():
                            class_start, class_end = row["Time"].split('-')
                            class_start = parse_time(class_start)
                            class_end = parse_time(class_end)
                            
                            if (class_start < slot_end and class_end > slot_start):
                                has_overlap = True
                                break
                        
                        if not has_overlap:
                            break_entries.append({
                                "Day": day,
                                "Time": f"{slot_start.strftime('%H:%M')}-{slot_end.strftime('%H:%M')}",
                                "Room": "N/A",
                                "Batch": batch,
                                "Course": "Lunch",
                                "Type": "BREAK",
                                "Faculty": "N/A",
                                "Details": "Lunch Break"
                            })
                            lunch_break_found = True
                            break
                
                if lunch_break_found:
                    break
            
            # And for afternoon break
            afternoon_break_found = False
            for start, end in afternoon_break_times:
                for slot in all_slots:
                    slot_start, slot_end = slot.split('-')
                    slot_start = parse_time(slot_start)
                    slot_end = parse_time(slot_end)
                    
                    if (start <= slot_start < end and slot_end <= end):
                        # Check if no classes overlap
                        has_overlap = False
                        for _, row in day_df.iterrows():
                            class_start, class_end = row["Time"].split('-')
                            class_start = parse_time(class_start)
                            class_end = parse_time(class_end)
                            
                            if (class_start < slot_end and class_end > slot_start):
                                has_overlap = True
                                break
                        
                        if not has_overlap:
                            break_entries.append({
                                "Day": day,
                                "Time": f"{slot_start.strftime('%H:%M')}-{slot_end.strftime('%H:%M')}",
                                "Room": "N/A",
                                "Batch": batch,
                                "Course": "Break",
                                "Type": "BREAK",
                                "Faculty": "N/A",
                                "Details": "Afternoon Break"
                            })
                            afternoon_break_found = True
                            break
                
                if afternoon_break_found:
                    break
                    
        return break_entries
    
    # Find empty periods (gaps between classes)
    def find_empty_periods(batch_df, all_slots):
        empty_entries = []
        days = batch_df["Day"].unique()
        batch = batch_df["Batch"].iloc[0] if not batch_df.empty else "UNKNOWN"
        
        for day in days:
            day_df = batch_df[batch_df["Day"] == day]
            used_slots = set(day_df["Time"].unique())
            
            for slot in all_slots:
                if slot in used_slots:
                    continue
                
                # Check if this slot overlaps with any used slot
                slot_start, slot_end = slot.split('-')
                slot_start = parse_time(slot_start)
                slot_end = parse_time(slot_end)
                
                has_overlap = False
                for used_slot in used_slots:
                    used_start, used_end = used_slot.split('-')
                    used_start = parse_time(used_start)
                    used_end = parse_time(used_end)
                    
                    if (slot_start < used_end and slot_end > used_start):
                        has_overlap = True
                        break
                
                # If no overlap and not too short, add as empty period
                if not has_overlap and (slot_end - slot_start).total_seconds() / 60 >= 20:
                    # Check if this is already a break
                    is_break = False
                    for item in empty_entries:
                        if item["Day"] == day and item["Time"] == slot:
                            is_break = True
                            break
                    
                    if not is_break:
                        empty_entries.append({
                            "Day": day,
                            "Time": slot,
                            "Room": "N/A",
                            "Batch": batch,
                            "Course": "Free",
                            "Type": "BREAK",
                            "Faculty": "N/A",
                            "Details": "Free Period"
                        })
        
        return empty_entries
    
    # Process each batch separately
    batches = sorted(df["Batch"].unique())
    
    # Define colors for different types of courses
    course_colors = {}
    
    for batch in batches:
        # Skip 'ALL' batch if processing individually
        if batch == "ALL":
            continue
            
        # Filter for this batch
        batch_df = df[df["Batch"] == batch]
        
        # Generate all possible time slots for this batch
        time_points = get_time_points_for_batch(batch_df)
        all_slots = generate_time_slots(time_points)
        
        # Add breaks and free periods
        break_entries = add_breaks_for_batch(batch_df, all_slots)
        empty_entries = find_empty_periods(batch_df, all_slots)
        
        # Combine with original data
        combined_entries = pd.concat([
            batch_df,
            pd.DataFrame(break_entries),
            pd.DataFrame(empty_entries)
        ], ignore_index=True)
        
        # Days order (abbreviated to match the example image)
        days_order = {"Monday": "MON", "Tuesday": "TUE", "Wednesday": "WED", 
                     "Thursday": "THU", "Friday": "FRI", "Saturday": "SAT"}
        
        # Add short day names
        combined_entries["ShortDay"] = combined_entries["Day"].map(days_order)
        
        # Get classroom for this batch (most common room)
        classroom = batch_df["Room"].mode().iloc[0] if not batch_df.empty else "TBD"
        
        # Format batch name for display (e.g., "CSE A 2023" from "CSE_A_2023")
        formatted_batch = batch.replace("_", " ")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = str(batch)
        
        # Add headers
        ws.merge_cells('A1:J1')
        ws['A1'] = institution_name
        ws['A1'].font = Font(bold=True, size=14, color="000080")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Time Table for {academic_session}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A3:J3')
        ws['A3'] = f"Batch: {formatted_batch}, Class Room: {classroom}"
        ws['A3'].font = Font(bold=True, size=11)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Set up column headers - Time row
        ws.row_dimensions[5].height = 25
        ws['A5'] = "Time"
        ws['A5'].fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
        ws['A5'].font = Font(bold=True)
        
        # Day column
        ws['A6'] = "Day"
        ws['A6'].fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
        ws['A6'].font = Font(bold=True)
        
        # Get all time slots for this batch, sorted chronologically
        batch_time_slots = sorted(combined_entries["Time"].unique(), 
                                 key=lambda x: parse_time(x.split('-')[0]))
        
        # Add time slots in header
        for i, time_slot in enumerate(batch_time_slots, start=1):
            col = get_column_letter(i + 1)
            ws[f'{col}5'] = time_slot
            ws[f'{col}5'].fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
            ws[f'{col}5'].font = Font(bold=True)
            ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column width based on duration
            try:
                start, end = time_slot.split('-')
                start_time = parse_time(start)
                end_time = parse_time(end)
                duration = (end_time - start_time).total_seconds() / 60
                
                # Scale width based on duration
                width = max(12, min(25, 12 + duration / 15))
                ws.column_dimensions[col].width = width
            except:
                ws.column_dimensions[col].width = 15
        
        # Add days and populate timetable
        short_days = ["MON", "TUE", "WED", "THU", "FRI"] 
        
        for i, day_abbr in enumerate(short_days, start=0):
            row = i + 7
            ws[f'A{row}'] = day_abbr
            ws[f'A{row}'].fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
            ws[f'A{row}'].font = Font(bold=True)
            
            # Get full day name
            full_day = next((d for d, abbr in days_order.items() if abbr == day_abbr), None)
            
            if full_day:
                # Filter for this day
                day_data = combined_entries[combined_entries["Day"] == full_day]
                
                # Fill in the timetable
                for time_idx, time_slot in enumerate(batch_time_slots, start=1):
                    col = get_column_letter(time_idx + 1)
                    
                    # Find data for this time slot and day
                    slot_data = day_data[day_data["Time"] == time_slot]
                    
                    if not slot_data.empty:
                        entry = slot_data.iloc[0]
                        course = entry["Course"]
                        
                        # Determine display text and color
                        if course in ["Break", "Lunch", "Free"]:
                            # For breaks
                            display_text = entry["Details"]
                            cell_color = "DDDDDD"  # Gray for breaks
                        else:
                            # For regular courses
                            if "Course_Name" in entry:
                                course_name = entry["Course_Name"]
                            else:
                                course_name = courses_info.get(course, {}).get("name", course)
                                
                            room = entry["Room"]
                            faculty = entry["Faculty"]
                            course_type = entry["Type"]
                            
                            display_text = f"{course}\n{course_type}\nRoom: {room}\n{faculty}"
                            
                            # Color the cell based on course
                            if course not in course_colors:
                                # Generate color options
                                colors = ['FFCCFF', 'CCFFCC', '9999FF', 'FFFF99', 
                                         'FF9999', '99FFFF', 'FFCC99', 'CC99FF']
                                course_colors[course] = colors[len(course_colors) % len(colors)]
                            
                            cell_color = course_colors[course]
                        
                        # Set cell value and format
                        ws[f'{col}{row}'] = display_text
                        ws[f'{col}{row}'].fill = PatternFill(start_color=cell_color, 
                                                           end_color=cell_color, 
                                                           fill_type="solid")
                    
                    # Apply formatting even if empty
                    if ws[f'{col}{row}'].value is not None:
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', 
                                                              vertical='center', 
                                                              wrap_text=True)
                        
                    # Add borders to all cells in timetable
                    ws[f'{col}{row}'].border = Border(left=Side(style='thin'), 
                                                    right=Side(style='thin'),
                                                    top=Side(style='thin'), 
                                                    bottom=Side(style='thin'))
        
        # Add course information table
        course_row = len(short_days) + 9  # Leave space after timetable
        
        # Add table header
        ws[f'A{course_row}'] = "Sl.No."
        ws[f'B{course_row}'] = "Course Code"
        ws[f'C{course_row}'] = "Course Title"
        ws[f'D{course_row}'] = "Credits (L-T-P-C)"
        ws[f'E{course_row}'] = "Faculty"
        
        # Style header
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{course_row}'].fill = PatternFill(start_color="BBBBBB", 
                                                      end_color="BBBBBB", 
                                                      fill_type="solid")
            ws[f'{col}{course_row}'].font = Font(bold=True)
            ws[f'{col}{course_row}'].border = Border(left=Side(style='thin'), 
                                                   right=Side(style='thin'),
                                                   top=Side(style='thin'), 
                                                   bottom=Side(style='thin'))
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        
        # Get unique courses
        unique_courses = batch_df["Course"].unique()
        
        # Populate course table
        for i, course in enumerate(unique_courses, start=1):
            if course in ["Break", "Lunch", "Free"]:
                continue
                
            current_row = course_row + i
            
            # Get course information
            course_data = batch_df[batch_df["Course"] == course].iloc[0]
            faculty = course_data["Faculty"]
            
            # Get course name and credits
            if "Course_Name" in course_data:
                course_name = course_data["Course_Name"]
            else:
                course_name = courses_info.get(course, {}).get("name", course)
                
            credits = courses_info.get(course, {}).get("credits", "3-0-0-3")
            
            # Add to table
            ws[f'A{current_row}'] = i
            ws[f'B{current_row}'] = course
            ws[f'C{current_row}'] = course_name
            ws[f'D{current_row}'] = credits
            ws[f'E{current_row}'] = faculty
            
            # Color the course code cell with matching color from timetable
            if course in course_colors:
                ws[f'B{current_row}'].fill = PatternFill(start_color=course_colors[course], 
                                                       end_color=course_colors[course], 
                                                       fill_type="solid")
            
            # Add borders
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{current_row}'].border = Border(left=Side(style='thin'), 
                                                        right=Side(style='thin'),
                                                        top=Side(style='thin'), 
                                                        bottom=Side(style='thin'))
        
        # Save workbook
        batch_file = f"{batch}_Timetable.xlsx"
        wb.save(batch_file)
        print(f"✅ Timetable saved for {batch} as {batch_file}")

# Example usage
if __name__ == "__main__":
    process_timetable("final_timetable.csv")
